import cv2
import numpy as np
from collections import Counter
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill
from openpyxl.styles import Border, Side
from openpyxl.styles import Alignment


def dominant_color(cell):
    # A kép átalakítása 2D-s tömbbé, ahol minden sor egy képpont
    data = np.reshape(cell, (-1, 3))
    
    # Kizárjuk a fekete színt a számításból
    # Fekete szín: [0, 0, 0], az összehasonlítás egy kis küszöbértékkel történik, ha zajos a kép
    non_black = np.all(data > [30, 30, 30], axis=1)  # Ahol az RGB értékek minden csatornán nagyobbak, mint 10
    data = data[non_black]
    
    # Ellenőrizzük, van-e még elegendő adat
    if data.size == 0:
        return None  # Nincs nem fekete szín
    
    # Leggyakoribb szín keresése azok között, amelyek nem feketék
    colors, count = np.unique(data, axis=0, return_counts=True)
    dominant = colors[count.argmax()]
    return dominant

def adjust_gamma(image, gamma=1.5):  # gamma < 1 sötétíti, gamma > 1 világosítja a képet
    invGamma = 1.0 / gamma
    table = np.array([((i / 255.0) ** invGamma) * 255 for i in np.arange(0, 256)]).astype("uint8")
    return cv2.LUT(image, table)


def find_closest_color(color, colors):
    min_distance = float('inf')
    closest_color_name = None
    for name, value in colors.items():
        distance = np.linalg.norm(color - value)
        if distance < min_distance:
            min_distance = distance
            closest_color_name = name
    return closest_color_name

def rgb_to_hex(rgb):
    # Az rgb egy lista vagy tuple három decimális értékkel: [67, 154, 101]
    return '{:02X}{:02X}{:02X}'.format(rgb[0], rgb[1], rgb[2])


minimum_area = 100


# Definiáljuk a 6 szín RGB kódját
colors = {
    'red': np.array([173, 44, 72]),
    'green': np.array([38, 145, 92]),
    'blue': np.array([16, 103, 168]),
    'yellow': np.array([204, 193, 44]),
    'orange': np.array([197, 129, 18]),
    'white': np.array([216, 204, 200])
}

# 6x9-es tömb létrehozása, ahol minden elem egy 3 elemű RGB színkód (kezdetben 0-val inicializálva)
color_array = np.zeros((6, 9, 3), dtype=np.uint8)  #színkódok tárolása
#ez meg a párja, itt a felismert színeket tárolom
color_array_finds = np.empty((6, 9), dtype=object) #felismert színek tárolása


mappa = "proba"
go = 1
flag = 0
gammavalue=2.5

log_path = str(mappa)+"/log.txt"

while go:
    print(f"\ngamma: {gammavalue}\n")
    for k in range (0,6):
        imgneve = str(mappa)+"/"+str(k+1)+".jpg"

        image = cv2.imread(imgneve)
        pirosdb = 0
        zolddb = 0
        kekdb = 0
        citromdb = 0
        narancsdb = 0
        feherdb = 0

        #cv2.imshow('eredeti', image)
        #cv2.waitKey(0)
        #cv2.destroyAllWindows()


        #90%os fekete megtartása, minden más fehér - > nem jött be, most 75% -> majdnem jó, 50% .... nehéz olyat találni ami minden mintára megfelelő (itt már inverz)
        gray_image = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
        _, thresholded_image = cv2.threshold(gray_image, 50, 255, cv2.THRESH_BINARY_INV)
        #bw_image = cv2.bitwise_not(thresholded_image)
        #cv2.imshow('Csak 50% fekete', thresholded_image)
        #cv2.waitKey(0)
        #cv2.destroyAllWindows()

        contours, _ = cv2.findContours(thresholded_image, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)

        #nem akarom rárajzolni a kész képre a kontúrokat
        #for contour in contours: 
        #    cv2.drawContours(image, [contour], -1, (0, 255, 0), 2)

        # Megjelenítés
        #cv2.imshow('Fekete területek kijelölve', image)
        #cv2.waitKey(0)
        #cv2.destroyAllWindows()    

        if contours:
            largest_contour = max(contours, key=cv2.contourArea)
        else:
            largest_contour = None

        # Bounding box meghatározása
        x, y, w, h = cv2.boundingRect(largest_contour)

        # Négyzetes bounding box kiszámítása
        side_length = max(w, h)
        center_x, center_y = x + w // 2, y + h // 2

        # Négyzetes bounding box sarkainak kiszámítása
        top_left_x = center_x - side_length // 2
        top_left_y = center_y - side_length // 2
        bottom_right_x = top_left_x + side_length
        bottom_right_y = top_left_y + side_length

        # Négyzet kirajzolása
        #cv2.rectangle(image, (top_left_x, top_left_y), (bottom_right_x, bottom_right_y), (0, 255, 0), 2)

        # Kép megjelenítése
        #cv2.imshow('Bounding Square', image)
        #cv2.waitKey(0)
        #cv2.destroyAllWindows()

        # Kivágjuk a bounding box által meghatározott részt
        cropped_image = image[y:y+h, x:x+w]

        # Eredeti kép, bounding box és kivágott kép megjelenítése
        #cv2.rectangle(image, (x, y), (x+w, y+h), (0, 255, 0), 2)  # Bounding box kirajzolása az eredeti képre
        #cv2.imshow('Original Image with Bounding Box', image)
        #cv2.imshow('Cropped Image', cropped_image)
        #cv2.waitKey(0)
        #cv2.destroyAllWindows()

        #mivel már nem a 9 négyzet színeinek átlagát veszem, hanem egy megadott területből veszek mintát, fölöslegessé vált a maszkolás, így kivettem ezeket a lépéseket


        gamma_corrected = adjust_gamma(cropped_image, gammavalue)

        #cv2.imshow('Gamma Corrected', gamma_corrected)
        #cv2.waitKey(0)
        #cv2.destroyAllWindows()
        image = gamma_corrected # ezzel a szerkesztett képpel folytatom a feldolgozást

        # Kép dimenzióinak lekérdezése
        height, width = image.shape[:2]

        # Vízszintes vonalak
        for i in range(1, 3):
            cv2.line(image, (0, i * height // 3), (width, i * height // 3), (0, 255, 0), thickness=2)

        # Függőleges vonalak
        for i in range(1, 3):
            cv2.line(image, (i * width // 3, 0), (i * width // 3, height), (0, 255, 0), thickness=2)

        # Kép megjelenítése
        #cv2.imshow('Divided Image', image)
        #cv2.waitKey(0)
        #cv2.destroyAllWindows()

        # Kép mentése
        #cv2.imwrite('divided_image.jpg', image)


        cell_height = height // 3
        cell_width = width // 3
        margin_height = int(cell_height * 0.33) 
        margin_width = int(cell_width * 0.33)

        # Minden cella középső részének kivágása és feldolgozása
        cells = []
        for i in range(3):  # Sorok
            for j in range(3):  # Oszlopok
                cell = image[i*cell_height:(i+1)*cell_height, j*cell_width:(j+1)*cell_width]
                cells.append(cell)

        #for cell in cells:
        #    cv2.imshow('cella', cell)
        #    cv2.waitKey(0)
        #    cv2.destroyAllWindows()



        for i in range(0,9):
            #if (i==3) or (i==6):
            #    print("\n\n")

            #print(i+1)

            image = cells[i]
            namestr= "cella"+str(i)
            #cv2.imshow(namestr,image)


            height, width = image.shape[:2]

            x = int(width/5)
            y = int(height/5)
            w = int(width/4)
            h = int(height/4)

            cropped_image = image[y:y+h, x:x+w]
            #cv2.imshow(f'Cropped Image{i}', cropped_image)
            #cv2.waitKey(0)
            #cv2.destroyAllWindows()

            image = cropped_image
            # Átlagos BGR szín kiszámítása
            average_color_per_row = np.average(image, axis=0)
            average_color = np.average(average_color_per_row, axis=0)
            average_color = np.uint8(average_color)  # Konvertálás egész színekre
            #print("Átlagos szín (BGR):", average_color)
            #print("B: ",average_color[0],"G: ",average_color[1],"R: ",average_color[2])
            b=average_color[0]
            g=average_color[1]
            r=average_color[2]

            #átlagos szín letárolása a tömbben RGB sorrendben:
            color_array[k,i] = [r,g,b]


            # Átlagos szín megjelenítése egy képen
            #average_image = np.zeros_like(image, np.uint8)
            #average_image[:] = average_color

            #cv2.imshow(f'Average Color{i}', average_image)


            color_to_find = np.array([r, g, b])
            # Legközelebbi szín megtalálása
            closest_color = find_closest_color(color_to_find, colors)
            #print(f"RGB The closest color to {color_to_find} is {closest_color}")
            color_array_finds[k,i] = closest_color






    #cv2.waitKey(0)
    #cv2.destroyAllWindows()

    stored_colors = {}

    # Tömb összes elemének megjelenítése (opcionális)
    print("Teljes szín tömb:")
    print(color_array)
    print(color_array_finds)

    #ez egy ellenőrzés, minden színből 9-et kell találni
    # Színek megszámolása
    color_counts = Counter(color_array_finds.flatten())
    print("Színek száma:")
    for color, count in color_counts.items():
        print(f"{color}: {count}")
        stored_colors[color] = count
    print("*****************************\n")
    print(stored_colors)
    print("*****************************\n")
    #közepek egyediségének ellenőrzése
    fifth_elements = color_array_finds[:, 4]
    unique_elements = np.unique(fifth_elements)
    is_unique = len(unique_elements) == len(fifth_elements)

    print("Az 5. elemek minden oszlopból:", fifth_elements)
    #print("Egyedi elemek:", unique_elements)
    print("Minden szín csak egyszer szerepel:", is_unique)

    # Ellenőrzés, hogy minden érték 9-e
    all_nines = all(value == 9 for value in stored_colors.values())

    # Eredmény kiíratása
    print(all_nines)  # True, ha minden érték 9, különben False
    
    if all_nines or flag:
        go=0
    else:
        if gammavalue < 0.1:
            gammavalue = 1
            flag = 1
        else:    
            gammavalue = gammavalue - 0.1
    


#log fájl írása
with open(log_path, 'w') as file:
    file.write("Teljes szín tömb:\n")
    file.write(str(color_array))
    file.write("\n")
    file.write(str(color_array_finds))
    file.write("\n")
    file.write("Színek száma:\n")
    for color, count in color_counts.items():
        file.write(f"{color}: {count}")
    file.write("\n")
    file.write(f"Az 5. elemek minden oszlopból: {fifth_elements}\n")
    file.write(f"Minden szín csak egyszer szerepel: {is_unique}")
    file.write(f"\nGamma értéke: {gammavalue}")




# Excel munkafüzet létrehozása
wb = Workbook()
ws = wb.active

# Vékony fekete keret
thin_border = Border(
    left=Side(style='thin', color='000000'),
    right=Side(style='thin', color='000000'),
    top=Side(style='thin', color='000000'),
    bottom=Side(style='thin', color='000000')
)

# Beállítjuk az első 12 sor magasságát 50 pixelre
for i in range(1, 13):
    ws.row_dimensions[i].height = 56.25

# Beállítjuk az első 9 oszlop szélességét 50 pixelre
for i in range(1, 10):
    col_letter = get_column_letter(i)
    ws.column_dimensions[col_letter].width = 10

offsets=[
    [4,4],
    [4,1],
    [1,4],
    [7,4],
    [4,7],
    [4,10]
    ]

#if is_unique:
for k in range (0,6):
    #1-es oldal:
    x=offsets[k][0] #oszlop offset
    y=offsets[k][1] #sor offset
    db=0
    for i in range (0,3):
        for j in range (0,3):
            rgb= color_array[k,db]
            hex_color = rgb_to_hex(rgb)
            fill_color = PatternFill(start_color=hex_color, end_color=hex_color, fill_type='solid')
            cell = ws.cell(row = i+y, column=j+x)
            cell.fill = fill_color
            cell.border = thin_border
            db+=1
for k in range (0,6):
    #1-es oldal:
    x=offsets[k][0]+11 #oszlop offset
    y=offsets[k][1] #sor offset
    db=0
    for i in range (0,3):
        for j in range (0,3):
            cell = ws.cell(row = i+y, column=j+x)
            cell.value = color_array_finds[k,db]
            cell.border = thin_border
            db+=1

for k in range (0,6):
    #1-es oldal:
    x=offsets[k][0]+22 #oszlop offset
    y=offsets[k][1] #sor offset
    db=0
    for i in range (0,3):
        for j in range (0,3):
            cell = ws.cell(row = i+y, column=j+x)
            temp = "R "+str(color_array[k][db][0])+"\n"+"G "+str(color_array[k][db][1])+"\n"+"B "+str(color_array[k][db][2])+"\n"
            cell.value = temp
            cell.border = thin_border
            cell.alignment = Alignment(wrap_text=True)
            db+=1

x=0
y=14

for color, count in color_counts.items():
    cell = ws.cell(row = y, column = 1)
    temp=str(f"{color}: {count}")
    cell.value = temp
    y+=1

cell = ws.cell(row = 21, column = 1)
temp=str(f"gamma: {gammavalue}")
cell.value = temp
            

# Eredmények mentése Excel fájlba
savepath = str(mappa)+"/results"+str(mappa)+".xlsx"
wb.save(savepath)
#print(f"\ngamma: {gammavalue}\n")
print("kész")
